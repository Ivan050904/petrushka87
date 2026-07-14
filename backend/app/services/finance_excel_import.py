from __future__ import annotations

import hashlib
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

FinanceSide = Literal["expense", "income"]

MONTH_SHEET_TO_PERIOD: dict[str, tuple[int, int]] = {
    "Апрель": (2025, 4),
    "май": (2025, 5),
    "Июнь": (2025, 6),
    "Июль": (2025, 7),
    "Август": (2025, 8),
    "Сентябрь": (2025, 9),
    "Октябрь": (2025, 10),
    "Ноябрь": (2025, 11),
    "Декабрь": (2025, 12),
    "Январь": (2026, 1),
    "Февраль": (2026, 2),
    "Март": (2026, 3),
}

NUMERIC_SHEET_PATTERN = re.compile(r"^(\d{2})\.(\d{4})$")


@dataclass(frozen=True)
class ParsedExcelFinanceRow:
    sheet_name: str
    row_number: int
    side: FinanceSide
    category: str
    amount: float
    transaction_date: str
    external_id: str
    raw_description: str


@dataclass(frozen=True)
class ParsedExcelWorkbook:
    rows: list[ParsedExcelFinanceRow]
    categories: tuple[str, ...]
    expense_count: int
    income_count: int
    expense_total: float
    income_total: float
    sheet_count: int


def build_external_id(
    *,
    sheet_name: str,
    row_number: int,
    side: FinanceSide,
    category: str,
    amount: float,
) -> str:
    payload = f"{sheet_name}|{row_number}|{side}|{category.strip().lower()}|{amount:.2f}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def sheet_to_transaction_date(sheet_name: str) -> str:
    numeric_match = NUMERIC_SHEET_PATTERN.match(sheet_name.strip())
    if numeric_match:
        month = int(numeric_match.group(1))
        year = int(numeric_match.group(2))
    else:
        period = MONTH_SHEET_TO_PERIOD.get(sheet_name.strip())
        if period is None:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        year, month = period

    last_day = monthrange(year, month)[1]
    return date(year, month, last_day).isoformat()


def transaction_date_to_datetime(value: str) -> datetime:
    year, month, day = (int(part) for part in value.split("-"))
    return datetime(year, month, day, 12, 0, tzinfo=UTC)


def _normalize_category(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    category = value.strip()
    if not category:
        return None
    if category.isdigit():
        return None
    lowered = category.casefold()
    if lowered in {"расходы", "доходы", "все расходы", "все доходы"}:
        return None
    if "расход" in lowered or "доход" in lowered:
        return None
    return category[:160]


def _detect_side_columns(worksheet, *, side: FinanceSide) -> tuple[int, int]:
    marker = "расход" if side == "expense" else "доход"
    fallback = (1, 2) if side == "expense" else (3, 4)
    for col in range(1, worksheet.max_column + 1):
        value = worksheet.cell(1, col).value
        if isinstance(value, str) and marker in value.casefold():
            return col, col + 1
    return fallback


def _normalize_amount(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        amount = float(value)
        return amount if amount > 0 else None
    return None


def parse_finance_workbook(path: str | Path) -> ParsedExcelWorkbook:
    workbook = load_workbook(path, data_only=True)
    rows: list[ParsedExcelFinanceRow] = []
    categories: set[str] = set()
    expense_total = 0.0
    income_total = 0.0
    expense_count = 0
    income_count = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        transaction_date = sheet_to_transaction_date(sheet_name)
        expense_category_col, expense_amount_col = _detect_side_columns(worksheet, side="expense")
        income_category_col, income_amount_col = _detect_side_columns(worksheet, side="income")

        for row_number in range(2, worksheet.max_row + 1):
            expense_category = _normalize_category(
                worksheet.cell(row_number, expense_category_col).value
            )
            expense_amount = _normalize_amount(
                worksheet.cell(row_number, expense_amount_col).value
            )
            if expense_category and expense_amount is not None:
                rows.append(
                    _build_row(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        side="expense",
                        category=expense_category,
                        amount=expense_amount,
                        transaction_date=transaction_date,
                    )
                )
                categories.add(expense_category)
                expense_total += expense_amount
                expense_count += 1

            income_category = _normalize_category(
                worksheet.cell(row_number, income_category_col).value
            )
            income_amount = _normalize_amount(
                worksheet.cell(row_number, income_amount_col).value
            )
            if income_category and income_amount is not None:
                rows.append(
                    _build_row(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        side="income",
                        category=income_category,
                        amount=income_amount,
                        transaction_date=transaction_date,
                    )
                )
                categories.add(income_category)
                income_total += income_amount
                income_count += 1

    sorted_categories = tuple(sorted(categories, key=lambda item: item.casefold()))
    return ParsedExcelWorkbook(
        rows=rows,
        categories=sorted_categories,
        expense_count=expense_count,
        income_count=income_count,
        expense_total=expense_total,
        income_total=income_total,
        sheet_count=len(workbook.sheetnames),
    )


def _build_row(
    *,
    sheet_name: str,
    row_number: int,
    side: FinanceSide,
    category: str,
    amount: float,
    transaction_date: str,
) -> ParsedExcelFinanceRow:
    external_id = build_external_id(
        sheet_name=sheet_name,
        row_number=row_number,
        side=side,
        category=category,
        amount=amount,
    )
    return ParsedExcelFinanceRow(
        sheet_name=sheet_name,
        row_number=row_number,
        side=side,
        category=category,
        amount=amount,
        transaction_date=transaction_date,
        external_id=external_id,
        raw_description=f"{sheet_name} / {category}",
    )


def row_to_entry_payload(row: ParsedExcelFinanceRow) -> dict[str, object]:
    direction = row.side
    return {
        "type": "finance",
        "title": row.category,
        "content": row.category,
        "metadata": {
            "amount": row.amount,
            "direction": direction,
            "currency": "RUB",
            "description": row.category,
            "category": row.category,
            "transaction_date": row.transaction_date,
            "kind": direction,
            "external_id": row.external_id,
            "import_batch_id": "excel_import",
            "source": "excel_import",
        },
        "created_at": transaction_date_to_datetime(row.transaction_date),
    }
